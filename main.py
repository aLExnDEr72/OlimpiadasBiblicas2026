import os
import random
import sys
import subprocess
import threading
import time
from pathlib import Path
import tkinter as tk
from tkinter import messagebox

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None

BASE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "preguntas.xlsx")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
DATA_DIR = os.path.join(BASE_DIR, "data")
AUDIO_DIR = os.path.join(BASE_DIR, "audio")

BASE_W = 1536
BASE_H = 1024

COLS = {
    "NumeroPregunta": "numero",
    "Pregunta": "pregunta",
    "RespuestaCorrecta": "correcta",
    "RespuestaIncorrecta1": "inc1",
    "RespuestaIncorrecta2": "inc2",
    "RespuestaIncorrecta3": "inc3",
    "Dificultad": "dificultad",
    "ReferenciaBiblica": "referencia",
}

# Permite que el Excel definitivo use encabezados equivalentes,
# por ejemplo "Numero Pregunta" en lugar de "NumeroPregunta".
COL_ALIASES = {
    "NumeroPregunta": ["NumeroPregunta", "Numero Pregunta", "Número Pregunta", "Número de Pregunta", "Numero de Pregunta"],
    "Pregunta": ["Pregunta"],
    "RespuestaCorrecta": ["RespuestaCorrecta", "Respuesta Correcta"],
    "RespuestaIncorrecta1": ["RespuestaIncorrecta1", "Respuesta Incorrecta 1"],
    "RespuestaIncorrecta2": ["RespuestaIncorrecta2", "Respuesta Incorrecta 2"],
    "RespuestaIncorrecta3": ["RespuestaIncorrecta3", "Respuesta Incorrecta 3"],
    "Dificultad": ["Dificultad"],
    "ReferenciaBiblica": ["ReferenciaBiblica", "Referencia Biblica", "Referencia Bíblica"],
}


class AudioManager:
    """Administrador de audio compatible con macOS y Windows.

    - macOS: usa afplay, que viene incluido en el sistema.
    - Windows: usa PowerShell + Windows Media Player/WPF MediaPlayer.

    Controla correctamente la música de fondo para que no se mezclen
    INTRO e INTERFAZ, y detiene todo al cerrar la aplicación.
    """
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.platform = sys.platform
        self.enabled = self.platform.startswith("darwin") or self.platform.startswith("win")
        self.loop_thread = None
        self.loop_process = None
        self.loop_stop = threading.Event()
        self.loop_generation = 0
        self.current_loop = None
        self.effect_processes = []

    def path(self, filename):
        return os.path.abspath(os.path.join(self.base_dir, filename))

    def _popen_player(self, filename, volume=1.0):
        if not self.enabled:
            return None
        path = self.path(filename)
        if not os.path.exists(path):
            return None

        try:
            if self.platform.startswith("darwin"):
                return subprocess.Popen(
                    ["afplay", "-v", str(volume), path],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )

            if self.platform.startswith("win"):
                # Windows no tiene afplay. Usamos PowerShell con WPF MediaPlayer,
                # que reproduce MP3/M4A en Windows 10/11 sin instalar pygame.
                # Volume acepta 0.0 a 1.0. Los archivos que necesitan más volumen
                # se deben dejar preamplificados físicamente.
                vol = max(0.0, min(float(volume), 1.0))
                uri = Path(path).as_uri().replace("'", "''")
                ps = (
                    "Add-Type -AssemblyName PresentationCore; "
                    f"$p = New-Object System.Windows.Media.MediaPlayer; "
                    f"$p.Open([Uri]'{uri}'); "
                    f"$p.Volume = {vol}; "
                    "$p.Play(); "
                    "Start-Sleep -Milliseconds 250; "
                    "while ($p.NaturalDuration.HasTimeSpan -eq $false) { Start-Sleep -Milliseconds 100 }; "
                    "$ms = [int]$p.NaturalDuration.TimeSpan.TotalMilliseconds; "
                    "Start-Sleep -Milliseconds ($ms + 250); "
                    "$p.Stop(); $p.Close();"
                )
                creationflags = 0
                if hasattr(subprocess, "CREATE_NO_WINDOW"):
                    creationflags = subprocess.CREATE_NO_WINDOW
                return subprocess.Popen(
                    ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-WindowStyle", "Hidden", "-Command", ps],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    creationflags=creationflags,
                )
        except Exception:
            return None

        return None

    def _cleanup_effects(self):
        vivos = []
        for proc in self.effect_processes:
            try:
                if proc is not None and proc.poll() is None:
                    vivos.append(proc)
            except Exception:
                pass
        self.effect_processes = vivos

    def play_effect(self, filename, volume=1.0):
        proc = self._popen_player(filename, volume)
        if proc is not None:
            self.effect_processes.append(proc)
            self._cleanup_effects()

    def start_loop(self, filename, volume=1.0):
        # Evita relanzar el mismo loop si ya está activo.
        if self.current_loop == filename and self.loop_thread and self.loop_thread.is_alive():
            return

        self.stop_loop(wait=True)
        self.current_loop = filename
        self.loop_generation += 1
        generation = self.loop_generation
        stop_event = threading.Event()
        self.loop_stop = stop_event

        def runner():
            while not stop_event.is_set() and generation == self.loop_generation:
                proc = self._popen_player(filename, volume)
                if generation != self.loop_generation or stop_event.is_set():
                    if proc is not None and proc.poll() is None:
                        try:
                            proc.terminate()
                        except Exception:
                            pass
                    break
                self.loop_process = proc
                if proc is None:
                    break
                while proc.poll() is None and not stop_event.is_set() and generation == self.loop_generation:
                    time.sleep(0.1)
                if proc.poll() is None:
                    try:
                        proc.terminate()
                    except Exception:
                        pass
                time.sleep(0.08)
            if generation == self.loop_generation:
                self.loop_process = None

        self.loop_thread = threading.Thread(target=runner, daemon=True)
        self.loop_thread.start()

    def stop_loop(self, wait=False):
        self.loop_generation += 1
        self.loop_stop.set()
        proc = self.loop_process
        if proc is not None and proc.poll() is None:
            try:
                proc.terminate()
            except Exception:
                pass
        if wait and self.loop_thread and self.loop_thread.is_alive():
            try:
                self.loop_thread.join(timeout=0.5)
            except Exception:
                pass
        self.loop_process = None
        self.current_loop = None

    def stop_all(self):
        self.stop_loop(wait=True)
        for proc in list(self.effect_processes):
            try:
                if proc is not None and proc.poll() is None:
                    proc.terminate()
            except Exception:
                pass
        self.effect_processes = []

class OlimpiadasBiblicasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Olimpiadas Bíblicas 2026")
        self.root.geometry("1280x800")
        self.root.minsize(1000, 650)
        self.root.configure(bg="#050914")
        self.root.bind("<F11>", self.toggle_fullscreen)
        self.root.bind("<Escape>", self.exit_fullscreen)
        self.fullscreen = False

        self.canvas = tk.Canvas(root, bg="#050914", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<Configure>", self.on_resize)

        self.bg_inicio = self.load_image("pantalla_inicio.png", rgb=True)
        self.bg_juego = self.load_image("fondo_juego.png", rgb=True)
        self.bg_photo = None
        self.img_refs = []

        self.sprites = {}
        for name in [
            "comodin_5050_on", "comodin_5050_off",
            "comodin_ancianos_on", "comodin_ancianos_off",
            "comodin_escritos_on", "comodin_escritos_off",
            "logro_escudo_on", "logro_escudo_off",
            "logro_espada_on", "logro_espada_off",
            "logro_corona_on", "logro_corona_off",
            "temporizador_base",
        ]:
            self.sprites[name] = self.load_image(name + ".png", rgb=False)

        self.modo = "inicio"
        self.tribu = ""
        self.entrada_tribu = None
        self.preguntas = []
        self.ronda = []
        self.indice = 0
        self.puntos_ronda = 0
        self.opciones_actuales = []
        self.respuesta_correcta = ""
        self.ocultas_5050 = set()
        self.comodin_5050_usado = False
        self.comodin_ancianos_usado = False
        self.comodin_escritos_usado = False
        self.timer_activo = False
        self.segundos = 30
        self.timer_job = None
        # Marcador temporal: se mantiene solo mientras la app está abierta.
        # Al cerrar y abrir nuevamente, inicia vacío.
        self.marcador = {}
        self.audio = AudioManager(AUDIO_DIR)
        self.timer_sound_started = False
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.mostrar_inicio()

    def load_image(self, filename, rgb=False):
        if Image is None:
            return None
        path = os.path.join(ASSETS_DIR, filename)
        try:
            return Image.open(path).convert("RGB" if rgb else "RGBA")
        except Exception:
            return None

    def sx(self, x):
        return x * self.canvas.winfo_width() / BASE_W

    def sy(self, y):
        return y * self.canvas.winfo_height() / BASE_H

    def box(self, x1, y1, x2, y2):
        return (self.sx(x1), self.sy(y1), self.sx(x2), self.sy(y2))

    def font_size(self, size):
        scale = min(self.canvas.winfo_width() / BASE_W, self.canvas.winfo_height() / BASE_H)
        return max(8, int(size * scale))

    def limpiar(self):
        self.canvas.delete("all")
        self.img_refs.clear()
        for child in self.canvas.winfo_children():
            child.destroy()
        self.entrada_tribu = None

    def draw_background(self, image):
        w = max(1, self.canvas.winfo_width())
        h = max(1, self.canvas.winfo_height())
        if image and ImageTk:
            resized = image.resize((w, h), Image.Resampling.LANCZOS)
            self.bg_photo = ImageTk.PhotoImage(resized)
            self.canvas.create_image(0, 0, image=self.bg_photo, anchor="nw")
        else:
            self.canvas.create_rectangle(0, 0, w, h, fill="#050914", outline="")

    def draw_sprite(self, name, cx, cy, w, h, tag=None, cmd=None):
        img = self.sprites.get(name)
        if img is None or ImageTk is None:
            self.canvas.create_rectangle(*self.box(cx-w/2, cy-h/2, cx+w/2, cy+h/2), outline="#d6a743", width=2)
            return None
        px_w = max(1, int(self.sx(w)))
        px_h = max(1, int(self.sy(h)))
        resized = img.resize((px_w, px_h), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(resized)
        self.img_refs.append(photo)
        item = self.canvas.create_image(self.sx(cx), self.sy(cy), image=photo, anchor="center", tags=tag)
        if tag and cmd:
            self.canvas.tag_bind(tag, "<Button-1>", lambda e: cmd())
            self.canvas.tag_bind(tag, "<Enter>", lambda e: self.canvas.config(cursor="hand2"))
            self.canvas.tag_bind(tag, "<Leave>", lambda e: self.canvas.config(cursor=""))
        return item

    def draw_invisible_button(self, x1, y1, x2, y2, cmd, tag):
        self.canvas.create_rectangle(*self.box(x1, y1, x2, y2), fill="", outline="", tags=tag)
        self.canvas.tag_bind(tag, "<Button-1>", lambda e: cmd())
        self.canvas.tag_bind(tag, "<Enter>", lambda e: self.canvas.config(cursor="hand2"))
        self.canvas.tag_bind(tag, "<Leave>", lambda e: self.canvas.config(cursor=""))

    def on_resize(self, _=None):
        if self.modo == "inicio":
            self.mostrar_inicio(redraw=True)
        else:
            self.mostrar_juego(redraw=True)

    def mostrar_inicio(self, redraw=False):
        self.modo = "inicio"
        if self.timer_job:
            self.root.after_cancel(self.timer_job)
            self.timer_job = None
        self.limpiar()
        self.draw_background(self.bg_inicio)
        self.audio.start_loop("INTRO.m4a", volume=1.0)

        # Campo nativo único para escribir el nombre de la tribu.
        # La etiqueta "NOMBRE DE LA TRIBU" ya forma parte de la imagen de fondo.
        self.entrada_tribu = tk.Entry(
            self.canvas,
            justify="center",
            fg="white",
            bg="#0b1a33",
            insertbackground="white",
            font=("Georgia", self.font_size(32), "bold"),
            relief="flat",
            highlightthickness=1,
            highlightbackground="#ffffff",
            highlightcolor="#f8d27a",
        )
        self.canvas.create_window(self.sx(768), self.sy(676), window=self.entrada_tribu,
                                  width=self.sx(620), height=self.sy(58))
        self.entrada_tribu.delete(0, "end")

        self.draw_podio_campeones()

        self.draw_invisible_button(510, 800, 1025, 895, self.nueva_ronda, "nueva_ronda")
        self.canvas.create_text(self.sx(768), self.sy(965),
                                text="F11: pantalla completa   |   ESC: salir de pantalla completa",
                                fill="#e8d5a0", font=("Arial", self.font_size(12)))

    def cargar_preguntas_excel(self):
        if load_workbook is None:
            raise RuntimeError("Falta instalar openpyxl. Ejecuta: python3 -m pip install openpyxl pillow")
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}

        # Resolver encabezados usando alias. Esto evita errores si el Excel final
        # usa espacios o tildes en algunos títulos de columna.
        resolved_idx = {}
        faltantes = []
        for canonical in COLS:
            found = None
            for alias in COL_ALIASES.get(canonical, [canonical]):
                if alias in idx:
                    found = idx[alias]
                    break
            if found is None:
                faltantes.append(canonical)
            else:
                resolved_idx[canonical] = found

        if faltantes:
            raise RuntimeError("Faltan columnas en el Excel: " + ", ".join(faltantes))

        preguntas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[resolved_idx["Pregunta"]]:
                continue
            item = {v: row[resolved_idx[k]] for k, v in COLS.items()}
            item["dificultad"] = int(item["dificultad"])
            preguntas.append(item)
        return preguntas

    def nueva_ronda(self):
        if self.entrada_tribu is not None:
            self.tribu = self.entrada_tribu.get().strip().upper()
            if not self.tribu:
                messagebox.showwarning("Nombre de la tribu", "Escribe el nombre de la tribu antes de iniciar la ronda.")
                return
        try:
            self.preguntas = self.cargar_preguntas_excel()
            d1 = [p for p in self.preguntas if p["dificultad"] == 1]
            d2 = [p for p in self.preguntas if p["dificultad"] == 2]
            d3 = [p for p in self.preguntas if p["dificultad"] == 3]
            if len(d1) < 3 or len(d2) < 3 or len(d3) < 4:
                raise RuntimeError("La base debe tener mínimo 3 preguntas de dificultad 1, 3 de dificultad 2 y 4 de dificultad 3.")
            self.ronda = random.sample(d1, 3) + random.sample(d2, 3) + random.sample(d3, 4)
        except Exception as e:
            messagebox.showerror("Error al iniciar", str(e))
            return

        self.indice = 0
        self.puntos_ronda = 0
        self.opciones_actuales = []
        self.ocultas_5050 = set()
        self.comodin_5050_usado = False
        self.comodin_ancianos_usado = False
        self.comodin_escritos_usado = False
        self.timer_activo = False
        self.segundos = 30
        self.timer_sound_started = False
        self.audio.stop_loop()
        self.audio.play_effect("TRANSICION.m4a", volume=1.0)
        self.audio.start_loop("INTERFAZ.mp3", volume=0.5)
        self.mostrar_juego()

    def mostrar_juego(self, redraw=False):
        self.modo = "juego"
        self.limpiar()
        self.draw_background(self.bg_juego)
        self.draw_comodines()
        self.draw_logros()
        self.draw_timer_base()
        self.draw_textos_dinamicos()
        if self.ronda:
            self.dibujar_pregunta_y_respuestas()

    def draw_comodines(self):
        w, h = 330, 92
        self.draw_sprite("comodin_5050_off" if self.comodin_5050_usado else "comodin_5050_on", 190, 120, w, h, "comodin_5050", self.usar_5050)
        self.draw_sprite("comodin_ancianos_off" if self.comodin_ancianos_usado else "comodin_ancianos_on", 190, 245, w, h, "comodin_ancianos", self.usar_ancianos)
        self.draw_sprite("comodin_escritos_off" if self.comodin_escritos_usado else "comodin_escritos_on", 190, 370, w, h, "comodin_escritos", self.usar_escritos)

    def draw_logros(self):
        # Orden superior a inferior: Corona, espada, escudo.
        corona = "logro_corona_on" if self.puntos_ronda >= 200 else "logro_corona_off"
        espada = "logro_espada_on" if self.puntos_ronda >= 100 else "logro_espada_off"
        escudo = "logro_escudo_on" if self.puntos_ronda >= 50 else "logro_escudo_off"
        self.draw_sprite(corona, 1325, 120, 310, 105)
        self.draw_sprite(espada, 1325, 245, 310, 105)
        self.draw_sprite(escudo, 1325, 370, 310, 105)

    def draw_timer_base(self):
        # El temporizador base ya forma parte de la imagen de fondo.
        # Aquí solo se dibuja el número dinámico, con fondo transparente,
        # cuando el comodín Consultar los Escritos está activo.
        if self.timer_activo:
            color = "#fff1c2" if self.segundos > 10 else "#ff554d"
            self.canvas.create_text(self.sx(1407), self.sy(900), text=str(self.segundos), fill=color,
                                    font=("Arial", self.font_size(54), "bold"))
            self.draw_referencia_biblica()

    def draw_referencia_biblica(self):
        # Muestra la referencia bíblica solo mientras está activo el temporizador
        # del comodín Consultar los Escritos.
        if not self.timer_activo or not self.ronda:
            return
        try:
            referencia = str(self.ronda[self.indice].get("referencia", "")).strip()
        except Exception:
            referencia = ""
        if not referencia or referencia.lower() == "none":
            return

        # Ubicación: justo debajo del comodín Consultar los Escritos.
        # Se dibuja como texto dinámico con fondo transparente para no tapar la imagen.
        self.canvas.create_text(self.sx(190), self.sy(438), text=referencia.upper(),
                                fill="#f8d27a",
                                font=("Georgia", self.font_size(18), "bold"),
                                width=self.sx(320), justify="center")

    def nombre_tribu_para_juego(self):
        nombre = (self.tribu or "").strip().upper()
        if nombre.startswith("TRIBU DE "):
            return nombre
        if nombre.startswith("TRIBU "):
            return nombre
        return f"TRIBU DE {nombre}"

    def draw_textos_dinamicos(self):
        # Nombre de la tribu en recuadro vacío del fondo.
        self.canvas.create_text(self.sx(768), self.sy(455), text=self.nombre_tribu_para_juego(), fill="#f8d27a",
                                font=("Georgia", self.font_size(25), "bold"), width=self.sx(440), justify="center")
        # Número de pregunta dinámico bajo la palabra PREGUNTA.
        self.canvas.create_text(self.sx(768), self.sy(535), text=f"{self.indice + 1} DE 10", fill="#f8d27a",
                                font=("Georgia", self.font_size(17), "bold"), justify="center")

    def preparar_opciones(self):
        p = self.ronda[self.indice]
        self.respuesta_correcta = str(p["correcta"])
        opciones = [str(p["correcta"]), str(p["inc1"]), str(p["inc2"]), str(p["inc3"])]
        random.shuffle(opciones)
        self.opciones_actuales = opciones
        self.ocultas_5050 = set()

    def dibujar_pregunta_y_respuestas(self):
        if not self.opciones_actuales:
            self.preparar_opciones()
        p = self.ronda[self.indice]
        self.canvas.create_text(self.sx(768), self.sy(625), text=str(p["pregunta"]), fill="white",
                                font=("Arial", self.font_size(28), "bold"), width=self.sx(880), justify="center")
        coords = [(480, 782), (1060, 782), (480, 905), (1060, 905)]
        letras = ["A", "B", "C", "D"]
        for i, (cx, cy) in enumerate(coords):
            texto = "" if i in self.ocultas_5050 else self.opciones_actuales[i]
            self.canvas.create_text(self.sx(cx + 35), self.sy(cy), text=texto, fill="white",
                                    font=("Arial", self.font_size(23), "bold"), width=self.sx(355), justify="center")
            if texto:
                self.draw_invisible_button(cx - 250, cy - 42, cx + 250, cy + 42, lambda k=i: self.responder(k), f"ans_{i}")

    def mostrar_pregunta(self):
        self.opciones_actuales = []
        self.ocultas_5050 = set()
        self.timer_activo = False
        self.timer_sound_started = False
        if self.timer_job:
            self.root.after_cancel(self.timer_job)
            self.timer_job = None
        self.mostrar_juego()

    def responder(self, idx):
        if idx in self.ocultas_5050:
            return
        seleccion = self.opciones_actuales[idx]
        correcto = seleccion == self.respuesta_correcta
        self.feedback_respuesta(idx, correcto)
        if correcto:
            self.audio.play_effect("RESPUESTA CORRECTA.m4a", volume=1.0)
            self.root.after(800, self.respuesta_correcta_accion)
        else:
            # Al responder incorrectamente, se detiene de inmediato la música
            # de fondo de la interfaz para que el efecto dramático se escuche limpio.
            self.audio.stop_loop(wait=False)
            self.audio.play_effect("RESPUESTA INCORRECTA.m4a", volume=1.0)
            self.root.after(1000, self.respuesta_incorrecta_accion)

    def feedback_respuesta(self, idx, correcto):
        # Dibuja una luz simple sobre la respuesta elegida y sobre la correcta.
        coords = [(480, 782), (1060, 782), (480, 905), (1060, 905)]
        for i, (cx, cy) in enumerate(coords):
            if i in self.ocultas_5050:
                continue
            if self.opciones_actuales[i] == self.respuesta_correcta:
                fill = "#0b5c2b"
            elif i == idx and not correcto:
                fill = "#7a1111"
            else:
                continue
            self.canvas.create_rectangle(*self.box(cx - 250, cy - 42, cx + 250, cy + 42), fill=fill, outline="#d6a743", width=2, stipple="gray50")
            self.canvas.create_text(self.sx(cx + 35), self.sy(cy), text=self.opciones_actuales[i], fill="white",
                                    font=("Arial", self.font_size(23), "bold"), width=self.sx(355), justify="center")

    def respuesta_correcta_accion(self):
        pregunta_num = self.indice + 1
        if pregunta_num == 3:
            self.puntos_ronda = 50
            self.audio.play_effect("LOGROS.m4a", volume=1.0)
            self.mostrar_juego()
            messagebox.showinfo("¡Logro obtenido!", "🛡 ESCUDO OBTENIDO\n+50 pt")
        elif pregunta_num == 6:
            self.puntos_ronda = 100
            self.audio.play_effect("LOGROS.m4a", volume=1.0)
            self.mostrar_juego()
            messagebox.showinfo("¡Logro obtenido!", "⚔ ESPADA OBTENIDA\n+100 pt")
        elif pregunta_num == 10:
            self.puntos_ronda = 200
            self.audio.play_effect("LOGROS.m4a", volume=1.0)
            self.mostrar_juego()
            messagebox.showinfo("¡Victoria!", "👑 CORONA OBTENIDA\n+200 pt")
            self.finalizar_ronda()
            return
        else:
            messagebox.showinfo("Resultado", "¡RESPUESTA CORRECTA!")
        self.indice += 1
        if self.indice < len(self.ronda):
            self.mostrar_pregunta()
        else:
            self.finalizar_ronda()

    def respuesta_incorrecta_accion(self):
        messagebox.showinfo("Resultado", f"RESPUESTA INCORRECTA\n\nLa tribu conserva: {self.puntos_ronda} pt")
        self.finalizar_ronda()

    def finalizar_ronda(self):
        nombre = self.tribu.upper()
        self.marcador[nombre] = self.marcador.get(nombre, 0) + self.puntos_ronda
        tabla = "\n".join([f"{k}: {v} pt" for k, v in sorted(self.marcador.items(), key=lambda x: -x[1])])
        messagebox.showinfo("Ronda finalizada", f"Puntos ganados: {self.puntos_ronda} pt\n\nPODIO DE CAMPEONES\n{tabla}")
        self.opciones_actuales = []
        self.mostrar_inicio()

    def usar_5050(self):
        if self.comodin_5050_usado or not self.opciones_actuales:
            return
        self.comodin_5050_usado = True
        self.audio.play_effect("COMODIN.m4a", volume=1.0)
        incorrectas = [i for i, op in enumerate(self.opciones_actuales) if op != self.respuesta_correcta]
        self.ocultas_5050 = set(random.sample(incorrectas, 2))
        self.mostrar_juego()

    def usar_ancianos(self):
        if self.comodin_ancianos_usado:
            return
        self.comodin_ancianos_usado = True
        self.audio.play_effect("COMODIN.m4a", volume=1.0)
        self.mostrar_juego()
        messagebox.showinfo("Consejo de Ancianos", "La tribu puede consultar al público.\nLuego debe decidir si acepta o no la sugerencia.")

    def usar_escritos(self):
        if self.comodin_escritos_usado:
            return
        self.comodin_escritos_usado = True
        self.audio.play_effect("COMODIN.m4a", volume=1.0)
        self.timer_activo = True
        self.segundos = 30
        self.timer_sound_started = False
        self.mostrar_juego()
        self.actualizar_timer()

    def actualizar_timer(self):
        if not self.timer_activo:
            return
        if self.segundos == 10 and not self.timer_sound_started:
            self.timer_sound_started = True
            self.audio.play_effect("TEMPORIZADOR_10s.m4a", volume=1.0)
        self.mostrar_juego()
        if self.segundos <= 0:
            self.timer_activo = False
            self.timer_job = None
            messagebox.showinfo("Tiempo", "TIEMPO AGOTADO")
            self.mostrar_juego()
            return
        self.segundos -= 1
        self.timer_job = self.root.after(1000, self.actualizar_timer)

    def draw_podio_campeones(self):
        # Marcador general visible en pantalla de inicio.
        # No se guarda en disco: se reinicia cada vez que se abre la aplicación.
        # Panel 25% más angosto, con fondo semitransparente y un solo marco ocre.
        x1, y1, x2, y2 = 1188, 250, 1472, 585

        # Fondo azul semitransparente simulado con stipple gray50.
        # El texto se mantiene 100% opaco para buena lectura.
        self.canvas.create_rectangle(*self.box(x1, y1, x2, y2),
                                     fill="#071326", stipple="gray25",
                                     outline="#6f5422", width=max(1, self.font_size(2)))

        self.canvas.create_text(self.sx((x1 + x2) / 2), self.sy(y1 + 42),
                                text="PODIO DE CAMPEONES", fill="#f8d27a",
                                font=("Georgia", self.font_size(20), "bold"), justify="center")
        if not self.marcador:
            return
        ordenados = sorted(self.marcador.items(), key=lambda x: (-x[1], x[0]))
        y = y1 + 92
        medallas = ["🥇", "🥈", "🥉"]
        for i, (tribu, puntos) in enumerate(ordenados[:8], start=1):
            prefijo = medallas[i - 1] if i <= 3 else f"{i}."
            linea = f"{prefijo} {tribu}: {puntos} pt"
            self.canvas.create_text(self.sx(x1 + 22), self.sy(y), text=linea, anchor="w",
                                    fill="white", font=("Arial", self.font_size(18), "bold"),
                                    width=self.sx(x2 - x1 - 34))
            y += 34


    def on_close(self):
        if self.timer_job:
            try:
                self.root.after_cancel(self.timer_job)
            except Exception:
                pass
            self.timer_job = None
        self.audio.stop_all()
        self.root.destroy()

    def toggle_fullscreen(self, event=None):
        self.fullscreen = not self.fullscreen
        self.root.attributes("-fullscreen", self.fullscreen)

    def exit_fullscreen(self, event=None):
        self.fullscreen = False
        self.root.attributes("-fullscreen", False)

if __name__ == "__main__":
    root = tk.Tk()
    app = OlimpiadasBiblicasApp(root)
    root.mainloop()
